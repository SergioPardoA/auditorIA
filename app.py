import streamlit as st
import pandas as pd
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from datetime import datetime
import matplotlib.pyplot as plt
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as PDFTable, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo

# ---------- CONFIGURACIÓN VISUAL ----------
st.set_page_config(page_title="Auditoría Inteligente", layout="wide", page_icon="📊")
st.markdown("""
    <style>
    .main {
        background-color: #f5f7fb;
    }
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    .title {
        font-size: 2.5rem;
        color: #003366;
        text-align: center;
        font-weight: bold;
        margin-bottom: 1rem;
    }
    .subtitle {
        font-size: 1.2rem;
        color: #444;
        text-align: center;
        margin-bottom: 2rem;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="title">📊 Auditoría Inteligente con IA</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Analiza tus asientos contables, detecta anomalías y genera informes automáticamente.</div>', unsafe_allow_html=True)

# ---------- CSV DE EJEMPLO ----------
@st.cache_data
def descargar_csv_ejemplo():
    ejemplo = pd.DataFrame({
        'Fecha': ['2024-01-01', '2024-01-01', '2024-01-02', '2024-01-03'],
        'Cuenta': [7000, 7000, 4300, 1000],
        'Debe': [1000, 1000, 0, 1500],
        'Haber': [0, 0, 1000, 0],
        'Documento': ['INV001', 'INV001', 'INV002', 'COMPRA1'],
        'Hora': ['07:45', '07:45', '12:00', '20:00']
    })
    output = BytesIO()
    ejemplo.to_csv(output, index=False)
    return output.getvalue()

st.download_button(
    label="📥 Descargar CSV de ejemplo",
    data=descargar_csv_ejemplo(),
    file_name="ejemplo_asientos.csv",
    mime="text/csv",
    help="Archivo de ejemplo para probar la app"
)

# ---------- CARGA DE ARCHIVO ----------
archivo = st.file_uploader("📁 Sube tu archivo de asientos contables (CSV)", type=["csv"])

if archivo is not None:
    try:
        df = pd.read_csv(archivo)
        df.columns = df.columns.str.strip().str.lower()
        mapeo_columnas = {
            'fecha': 'Fecha', 'cuenta': 'Cuenta', 'debe': 'Debe', 'haber': 'Haber',
            'documento': 'Documento', 'referencia': 'Documento', 'factura': 'Documento',
            'hora': 'Hora', 'hora operacion': 'Hora', 'hora operación': 'Hora'
        }
        df.rename(columns=mapeo_columnas, inplace=True)

        columnas_obligatorias = {"Fecha", "Cuenta", "Debe", "Haber"}
        columnas_opcionales = {"Documento", "Hora"}
        presentes = set(df.columns)
        faltan_obligatorias = columnas_obligatorias - presentes
        faltan_opcionales = columnas_opcionales - presentes
        columnas_extra = presentes - (columnas_obligatorias | columnas_opcionales)

        if faltan_obligatorias:
            st.error(f"❌ El archivo no es válido. Faltan columnas obligatorias: {', '.join(faltan_obligatorias)}")
            st.stop()
        if faltan_opcionales:
            st.warning(f"⚠️ Faltan columnas opcionales: {', '.join(faltan_opcionales)}. Algunas funciones pueden no aplicarse.")
        if columnas_extra:
            st.info(f"ℹ️ Las siguientes columnas se ignorarán: {', '.join(columnas_extra)}")

        def normalizar_hora(df):
            hora_col = df["Hora"].astype(str).str.strip()
            def parse_hora(valor):
                if ":" in valor:
                    try: return pd.to_datetime(valor, format="%H:%M").hour
                    except: return None
                elif valor.isdigit() and len(valor) > 2:
                    try: return int(valor) // 100
                    except: return None
                else:
                    try: return int(valor)
                    except: return None
            df["Hora_Normalizada"] = hora_col.apply(parse_hora)
            return df

        def procesar_datos(df):
            df["Importe"] = df[["Debe", "Haber"]].max(axis=1)
            df["FueraHorario"] = df.get("Hora_Normalizada", pd.Series(False)) < 8
            df["FueraHorario"] |= df.get("Hora_Normalizada", pd.Series(False)) > 18
            df["Redondeado"] = df["Importe"].astype(str).str.endswith(("000", "99"))
            df["Duplicado"] = df.duplicated(subset=["Fecha", "Cuenta", "Importe", "Documento"], keep=False) if "Documento" in df.columns else False
            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
            df["Fecha_Num"] = (df["Fecha"] - df["Fecha"].min()).dt.days
            features = df[["Fecha_Num", "Cuenta", "Importe"]].dropna()
            scaled = StandardScaler().fit_transform(features)
            modelo = IsolationForest(contamination=0.05, random_state=42)
            outliers = modelo.fit_predict(scaled) == -1
            df["Outlier"] = False
            df.loc[features.index, "Outlier"] = outliers
            return df

        df = normalizar_hora(df)
        df = procesar_datos(df)
        st.success("✅ Análisis completado")

        st.subheader("📈 Visualización de Outliers")
        fig, ax = plt.subplots()
        ax.scatter(df["Fecha_Num"], df["Importe"], c=df["Outlier"].map({True: 'red', False: 'blue'}), alpha=0.6)
        ax.set_xlabel("Fecha (días desde inicio)")
        ax.set_ylabel("Importe")
        ax.set_title("Distribución de valores y detección de outliers")
        st.pyplot(fig)

        st.subheader("📋 Resultados de Auditoría")
        st.dataframe(df)

        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as PDFTable, TableStyle
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        def generar_pdf(df, columnas_opcionales_ausentes, columnas_ignoradas):
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='TitleClean', fontSize=18, alignment=1, spaceAfter=10, textColor=colors.HexColor('#003366')))
            styles.add(ParagraphStyle(name='SectionBold', fontSize=13, fontName='Helvetica-Bold', spaceAfter=8))
            styles.add(ParagraphStyle(name='NormalText', fontSize=10, leading=14))
            elementos = []
            total = len(df)
            elementos.append(Spacer(1, 20))
            elementos.append(Paragraph("Informe de Auditoría con IA", styles['TitleClean']))
            elementos.append(Spacer(1, 12))
            elementos.append(Paragraph(f"Fecha de procesamiento: {datetime.now().strftime('%d/%m/%Y')}", styles['NormalText']))
            elementos.append(Paragraph(f"Total de registros analizados: {total}", styles['NormalText']))
            elementos.append(Spacer(1, 16))
            elementos.append(Paragraph("Resumen de Anomalías Detectadas:", styles['SectionBold']))
            anomalias = {
                "Operaciones fuera de horario": df.get("FueraHorario", pd.Series(dtype=bool)).sum(),
                "Importes redondeados": df.get("Redondeado", pd.Series(dtype=bool)).sum(),
                "Duplicados": df.get("Duplicado", pd.Series(dtype=bool)).sum(),
                "Outliers (IA)": df.get("Outlier", pd.Series(dtype=bool)).sum()
            }
            tabla = [["Tipo de Anomalía", "Cantidad", "Porcentaje"]]
            for desc, cant in anomalias.items():
                tabla.append([desc, str(cant), f"{cant/total*100:.2f}%"])
            t = PDFTable(tabla, colWidths=[230, 90, 90])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C0C0C0')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
            ]))
            elementos.append(t)
            elementos.append(Spacer(1, 20))
            elementos.append(Paragraph("Información sobre las columnas procesadas:", styles['SectionBold']))
            if columnas_opcionales_ausentes:
                elementos.append(Paragraph(f"⚠️ No se incluyeron las siguientes columnas opcionales: {', '.join(columnas_opcionales_ausentes)}", styles['NormalText']))
            else:
                elementos.append(Paragraph("✅ Todas las columnas opcionales estaban presentes.", styles['NormalText']))
            if columnas_ignoradas:
                elementos.append(Paragraph(f"ℹ️ Se ignoraron las siguientes columnas adicionales: {', '.join(columnas_ignoradas)}", styles['NormalText']))
            else:
                elementos.append(Paragraph("✅ No se detectaron columnas adicionales innecesarias.", styles['NormalText']))
            elementos.append(Spacer(1, 20))
            elementos.append(Paragraph("Interpretación de Hallazgos:", styles['SectionBold']))
            texto = [
                "1. Revisar operaciones fuera del horario laboral.",
                "2. Verificar si los importes redondeados tienen justificación.",
                "3. Comprobar si los duplicados son errores.",
                "4. Validar outliers para detectar fraudes."
            ]
            for linea in texto:
                elementos.append(Paragraph(linea, styles['NormalText']))
            doc.build(elementos)
            buffer.seek(0)
            return buffer

        def generar_excel(df):
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
            ws.freeze_panes = "A2"
            table_range = f"A1:{chr(64 + df.shape[1])}{len(df) + 1}"
            excel_table = ExcelTable(displayName="TablaAsientos", ref=table_range)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            excel_table.tableStyleInfo = style
            ws.add_table(excel_table)
            wb.save(output)
            output.seek(0)
            return output

        st.subheader("📥 Descargas")
        pdf_bytes = generar_pdf(df, list(faltan_opcionales), list(columnas_extra))
        excel_bytes = generar_excel(df)
        st.download_button("📄 Descargar PDF del Informe", data=pdf_bytes, file_name="informe_auditoria.pdf")
        st.download_button("📊 Descargar Excel Analizado", data=excel_bytes, file_name="asientos_analizados.xlsx")

    except Exception as e:
        st.error(f"⚠️ Error inesperado: {e}")
